"""
06_wide_format_output.py
========================
Creates a clean wide-format dataset from the extracted data.
  - One row per paper
  - One column per variable (all extracted, not just ML features)
  - Inputs and outputs labelled
  - Sensory values normalised to /10 scale
  - Saved as Excel with a variable guide sheet
"""

import pandas as pd
import numpy as np
import os
import sys
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")

# ── Configuration ─────────────────────────────────────────────────────────────
EXTRACTED_XLSX = r"D:\Monisha\UCC Project\outputs\extended_with_ocr.xlsx"
WIDE_XLSX      = r"D:\Monisha\UCC Project\outputs\wide_format_dataset.xlsx"

# ── Variable classification ───────────────────────────────────────────────────
INPUT_VARS = [
    "Oil_Fat_type",
    "Oil_type",
    "Fat_concentration_wt%",
    "Oil_concentration_wt%",
    "Protein_type",
    "Protein_concentration_wt%",
    "Concentration_wt%",
    "Emulsion_system",
    "Temperature_C",
    "Homogenization_rpm",
    "Pressure_MPa",
    "Pressure_bar",
    "Volume_mL",
    "Shear_rate_s-1",
    "Processing_time_s",
    "Participants_n",
]

OUTPUT_VARS = [
    "Viscosity_Pa_s",
    "d90_µm",
    "d50_µm",
    "d43_µm",
    "d32_µm",
    "Particle_size_µm",
    "Sensory_thickness",
    "Sensory_creaminess",
    "Sensory_smoothness",
    "Friction_coefficient",
    "Lubrication_property",
    "Sliding_speed_mm_s",
    "Normal_force_N",
]

SENSORY_VARS = ["Sensory_thickness", "Sensory_creaminess", "Sensory_smoothness"]

VARIABLE_UNITS = {
    "Fat_concentration_wt%":     "wt%",
    "Oil_concentration_wt%":     "wt%",
    "Protein_concentration_wt%": "wt%",
    "Concentration_wt%":         "wt%",
    "Temperature_C":             "°C",
    "Homogenization_rpm":        "rpm",
    "Pressure_MPa":              "MPa",
    "Pressure_bar":              "bar",
    "Volume_mL":                 "mL",
    "Shear_rate_s-1":            "s⁻¹",
    "Processing_time_s":         "s",
    "Viscosity_Pa_s":            "Pa·s",
    "d90_µm":                    "µm",
    "d50_µm":                    "µm",
    "d43_µm":                    "µm",
    "d32_µm":                    "µm",
    "Particle_size_µm":          "µm",
    "Sensory_thickness":         "/10",
    "Sensory_creaminess":        "/10",
    "Sensory_smoothness":        "/10",
    "Friction_coefficient":      "—",
    "Sliding_speed_mm_s":        "mm/s",
    "Normal_force_N":            "N",
}

# ── Load and pivot ────────────────────────────────────────────────────────────
print("=" * 60)
print("WIDE-FORMAT DATASET GENERATOR")
print("=" * 60)

print("\nLoading extracted data...")
xl = pd.ExcelFile(EXTRACTED_XLSX)
frames = [pd.read_excel(xl, sheet_name=s) for s in xl.sheet_names]
all_df = pd.concat(frames, ignore_index=True)
all_df["Value"] = pd.to_numeric(all_df["Value"], errors="coerce")
all_df["Normalized_Variable"] = all_df["Normalized_Variable"].astype(str)

print(f"  Total rows:  {len(all_df):,}")
print(f"  Papers:      {all_df['Paper'].nunique()}")

# Pivot: one row per paper, median value per variable
wide = (
    all_df[all_df["Value"].notna()]
    .groupby(["Paper", "Normalized_Variable"])["Value"]
    .median()
    .unstack()
)
wide = wide.drop(columns=["nan"], errors="ignore")

# Sensory scale normalisation: /100 → /10
for col in SENSORY_VARS:
    if col in wide.columns:
        wide[col] = wide[col].apply(
            lambda v: round(v / 10, 3) if pd.notna(v) and v > 10 else v
        )

print(f"  Variables:   {wide.shape[1]}")
print(f"  All vars:    {sorted(wide.columns.tolist())}")

# ── Order columns: inputs first, outputs next, rest last ─────────────────────
available_inputs  = [v for v in INPUT_VARS  if v in wide.columns]
available_outputs = [v for v in OUTPUT_VARS if v in wide.columns]
other_vars        = [c for c in wide.columns
                     if c not in INPUT_VARS and c not in OUTPUT_VARS]

ordered_cols = available_inputs + available_outputs + other_vars
wide = wide[ordered_cols]

# ── Add Study_ID ──────────────────────────────────────────────────────────────
wide.index.name = "Paper"
wide = wide.reset_index()

def make_study_id(filename):
    """Short identifier from the paper filename."""
    name = (filename
            .replace("1-s2.0-", "")
            .replace("-mainext", "")
            .replace("-main", "")
            .replace("-v2", ""))
    # Elsevier PII → use the PII code
    if name.startswith("S") and len(name) > 8:
        return name[:14]
    # Others → first 22 chars
    return name[:22].strip("-_ ")

wide.insert(0, "Study_ID", wide["Paper"].apply(make_study_id))

# ── Row count column: how many values were extracted for this paper ───────────
row_counts = (
    all_df[all_df["Value"].notna()]
    .groupby("Paper")["Value"]
    .count()
    .rename("Rows_extracted")
)
wide = wide.merge(row_counts, on="Paper", how="left")

print(f"\n  Output shape: {wide.shape[0]} rows × {wide.shape[1]} columns")

# ── Save to Excel with colour formatting ──────────────────────────────────────
print("\nSaving...")

with pd.ExcelWriter(WIDE_XLSX, engine="openpyxl") as writer:

    # ── Sheet 1: Dataset ──────────────────────────────────────────────────────
    wide.to_excel(writer, sheet_name="Dataset", index=False)

    ws = writer.sheets["Dataset"]

    # Colour fills
    INPUT_FILL  = PatternFill("solid", fgColor="D9EAD3")   # green
    OUTPUT_FILL = PatternFill("solid", fgColor="CFE2F3")   # blue
    OTHER_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow
    HEADER_FONT = Font(bold=True)
    CENTER      = Alignment(horizontal="center")

    # Map column letter to type
    col_names = list(wide.columns)
    for col_idx, col_name in enumerate(col_names, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        if col_name in INPUT_VARS:
            cell.fill = INPUT_FILL
        elif col_name in OUTPUT_VARS:
            cell.fill = OUTPUT_FILL
        elif col_name not in ("Study_ID", "Paper", "Rows_extracted"):
            cell.fill = OTHER_FILL

    # Auto-width columns
    for col_idx, col_name in enumerate(col_names, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(str(col_name)),
                      *(len(str(ws.cell(row=r, column=col_idx).value or ""))
                        for r in range(2, ws.max_row + 1)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    # Freeze top row
    ws.freeze_panes = "C2"

    # ── Sheet 2: Variable Guide ───────────────────────────────────────────────
    guide_rows = []
    for v in available_inputs:
        guide_rows.append({
            "Variable":    v,
            "Type":        "Input (Formulation / Process)",
            "Unit":        VARIABLE_UNITS.get(v, "—"),
            "Papers_with_data": int(wide[v].notna().sum()),
        })
    for v in available_outputs:
        guide_rows.append({
            "Variable":    v,
            "Type":        "Output (Measured property)",
            "Unit":        VARIABLE_UNITS.get(v, "—"),
            "Papers_with_data": int(wide[v].notna().sum()),
        })
    for v in other_vars:
        guide_rows.append({
            "Variable":    v,
            "Type":        "Other",
            "Unit":        VARIABLE_UNITS.get(v, "—"),
            "Papers_with_data": int(wide[v].notna().sum()),
        })

    guide_df = pd.DataFrame(guide_rows)
    guide_df.to_excel(writer, sheet_name="Variable_Guide", index=False)

    ws2 = writer.sheets["Variable_Guide"]
    for col_idx in range(1, len(guide_df.columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws2.column_dimensions[col_letter].width = 35
    for cell in ws2[1]:
        cell.font = HEADER_FONT

    # ── Sheet 3: Coverage summary ─────────────────────────────────────────────
    coverage = pd.DataFrame({
        "Variable": ordered_cols,
        "Type": [
            "Input"  if c in INPUT_VARS  else
            "Output" if c in OUTPUT_VARS else "Other"
            for c in ordered_cols
        ],
        "Papers_with_data": [int(wide[c].notna().sum()) for c in ordered_cols],
        "Coverage_%":       [round(wide[c].notna().mean() * 100, 1) for c in ordered_cols],
        "Median_value":     [round(wide[c].median(), 4) if pd.api.types.is_numeric_dtype(wide[c]) else "—"
                             for c in ordered_cols],
    })
    coverage.to_excel(writer, sheet_name="Coverage", index=False)

print(f"  Saved: {WIDE_XLSX}")

# ── Print summary ─────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("COVERAGE SUMMARY")
print("=" * 60)
print(f"\n  {'Variable':<35} {'Type':<10} {'Papers':>7}  {'Cover%':>7}")
print("  " + "-" * 65)
for _, row in coverage.iterrows():
    print(f"  {row['Variable']:<35} {row['Type']:<10} {row['Papers_with_data']:>7}  {row['Coverage_%']:>6}%")

print(f"\n  Green  = Input variables")
print(f"  Blue   = Output variables")
print(f"  Yellow = Other extracted variables")
print("\nDone.")
